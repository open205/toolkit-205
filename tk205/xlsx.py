import openpyxl
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font
import os
import json
import re
import enum
import string
from schema205 import A205Schema
from schema205 import process_grid_set, unique_name_with_index

class SheetType(enum.Enum):
    FLAT = 0
    PERFORMANCE_MAP = 1
    ARRAY = 2


class A205XLSXNode:

    white_space_multiplier = 4

    cell_border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )

    unused_style = NamedStyle(name="Unused",
        fill=PatternFill(start_color='808B96', end_color='808B96', fill_type='solid')
    )

    tile_style = NamedStyle(name="Title",
        fill=PatternFill(start_color='00529B', end_color='00529B', fill_type='solid'),
        font=Font(color="FFFFFF", bold=True, sz=14)
    )

    heading_style = NamedStyle(name="Heading",
        fill=PatternFill(start_color='01AED8', end_color='01AED8', fill_type='solid'),
        font=Font(bold=True),
        border=cell_border
    )

    schema_style = NamedStyle(name="Schema",
        fill=PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        font=Font(bold=True),
        border=cell_border
    )

    grid_var_style = NamedStyle(name="Grid Variables",
        fill=PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        font=Font(bold=True, color='0070C0'),
        border=cell_border
    )


    value_style = NamedStyle(name="Value",
        fill=PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
        border=cell_border
    )

    def __init__(self, name, parent=None, tree=None, value=None, sheet_ref=None, option=None):
        self.children = []  # List of children A205XLSXNodes
        self.name = name  # Name of this node
        self.value = value  # Value (if any) of this node
        self.sheet_ref = sheet_ref  # Reference to another sheet (if applicable)
        self.parent = parent  # Parent A205XLSXNode of this node
        self.grid_set = None  # Ordered arrays of repeated grid variable values (used only for grid_variable nodes)

        if parent:
            # Inherit much information from parent
            self.lineage = self.parent.lineage + [name]  # List of parent node names (as strings)
            self.options = self.parent.options + [option]
            self.tree = self.parent.tree
            self.sheet = self.parent.child_sheet
            self.sheet_type = self.parent.child_sheet_type
            self.beg = self.parent.next_child_beg
            if self.sheet_type == SheetType.FLAT:
                self.next_child_beg = self.beg + 1
            else:
                self.next_child_beg = self.beg
        else:
            # Root node
            self.lineage = []
            self.options = []
            self.tree = tree
            self.sheet = self.tree.schema_type
            self.sheet_type = SheetType.FLAT
            self.child_sheet_type = self.sheet_type
            self.beg = 3
            self.next_child_beg = 3

        if self.sheet not in self.tree.sheets:
            self.tree.sheets.append(self.sheet)

        # These will be changed by any children
        self.end = self.beg

        self.increment_ancestors()

        self.child_sheet = self.sheet
        self.child_sheet_type = self.sheet_type

        # Initial detection of new sheets
        if self.sheet_ref:
            # Indicator of pointer to another sheet
            if "performance_map" in self.sheet_ref:
                self.child_sheet_type = SheetType.PERFORMANCE_MAP
            elif "_representation" in self.sheet_ref:
                self.child_sheet_type = SheetType.FLAT
            else:
                self.child_sheet_type = SheetType.ARRAY

            if self.child_sheet_type == SheetType.FLAT:
                self.next_child_beg = 3
            else:
                self.next_child_beg = 1

            self.child_sheet = self.sheet_ref

        if parent:
            self.parent.add_child(self)

    def add_child(self, node):
        '''
        Add a child node to this node.
        '''
        self.children.append(node)

    def get_num_ancestors(self):
        '''
        Count ancestors back to root.
        '''
        if self.parent:
            return len(self.parent.lineage)
        else:
            return 0

    def get_ancestor(self,generation):
        if generation == 0:
            return self
        else:
            return self.parent.get_ancestor(generation - 1)

    def increment_ancestors(self):
        '''
        Increment the begining and endings appropriately for ancestors
        '''
        if self.parent:
            if self.sheet == self.parent.sheet:
                # If parent is in the same sheet increment both
                self.parent.end = self.end
                self.parent.increment_ancestors()
            self.parent.next_child_beg = self.end + 1

    def add_grid_set(self, grid_set):
        '''
        Add a cartesian product grid set to this node.

        Only used for "grid_variables" nodes.
        '''
        if self.name != 'grid_variables':
            raise Exception(f"Cannot add a grid set to '{self.name}'. Grid sets can only be added to 'grid_variables' nodes.")
        self.grid_set = grid_set

    def get_schema_node(self):
        '''
        Search for schema content for this node
        '''
        return self.tree.schema.get_schema_node(self.lineage, self.options)

    def is_required(self):
        '''
        Check schema if this is a required node
        '''
        if self.parent:
            parent_schema_node = self.parent.get_schema_node()
            if 'required' in parent_schema_node:
                return self.name in parent_schema_node['required']
            else:
                return False
        else:
            return True

    def fill_sheet(self, worksheet):
        for column in [x for x in string.ascii_uppercase] + ["A" + x for x in string.ascii_uppercase]:
            worksheet.column_dimensions[column].fill = self.unused_style.fill
            worksheet[f"{column}1"].style = self.tile_style

    def write_header(self, worksheet):
        '''
        Write the header data for a new sheet
        '''
        self.fill_sheet(worksheet)
        if self.parent:
            if self.sheet == self.parent.sheet:
                # A simple array (not a nested object)
                worksheet.cell(row=1, column=1).value = '.'.join(self.lineage)
                return
            else:
                worksheet.cell(row=1, column=1).value = '.'.join(self.parent.lineage)
        else:
            worksheet.cell(row=1, column=1).value = f"{self.tree.schema_type}: {self.tree.schema.get_rs_title(self.tree.schema_type)}"
        if self.sheet_type == SheetType.FLAT:
            xlsx_headers = ['Data Group', 'Data Element', 'Value', 'Units', 'Required']
            for column, header in enumerate(xlsx_headers, start=1):
                worksheet.cell(row=2, column=column).value = header
                worksheet.cell(row=2, column=column).style = self.heading_style
            worksheet.column_dimensions['B'].width = 50
            worksheet.column_dimensions['C'].width = 31

    def write_node(self):
        '''
        Write tree content to XLSX
        '''
        wb = self.tree.workbook
        sheet = self.sheet
        if sheet not in wb:
            wb.create_sheet(sheet)
            self.write_header(wb[sheet])

        schema_node = self.get_schema_node()

        if self.name is not None:
            if self.sheet_type == SheetType.FLAT:

                if len(self.children) > 0:
                    level_index = 1
                    wb[sheet].cell(row=self.beg,column=level_index).value = '.'.join(self.lineage)
                else:
                    level_index = 2
                    buffer = ' '*(self.get_num_ancestors() - 1)*self.white_space_multiplier # TODO: get_num_ancestors_to_root for embedded RS sheets?
                    wb[sheet].cell(row=self.beg,column=level_index).value = buffer + self.name

                wb[sheet].cell(row=self.beg,column=1).style = self.schema_style
                wb[sheet].cell(row=self.beg,column=2).style = self.schema_style
                wb[sheet].cell(row=self.beg,column=4).style = self.schema_style
                wb[sheet].cell(row=self.beg,column=5).style = self.schema_style
                wb[sheet].cell(row=self.beg,column=3).style = self.value_style

                if schema_node:
                    # Add description
                    if 'description' in schema_node:
                        comment = openpyxl.comments.Comment(schema_node['description'],"ASHRAE 205")
                        wb[sheet].cell(row=self.beg,column=level_index).comment = comment

                    # Enum validation
                    if 'enum' in schema_node:
                        enumerants = f'"{",".join(schema_node["enum"])}"'
                        if len(enumerants) < 256: # Apparent limitation of written lists (TODO: https://stackoverflow.com/a/33532984/1344457)
                            dv = openpyxl.worksheet.datavalidation.DataValidation(type='list',formula1=enumerants,allow_blank=True)
                            wb[sheet].add_data_validation(dv)
                            dv.add(wb[sheet].cell(row=self.beg,column=3))

                    # Boolean validation
                    if 'type' in schema_node:
                        if schema_node['type'] == 'boolean':
                            dv = openpyxl.worksheet.datavalidation.DataValidation(type='list',formula1='"TRUE,FALSE"',allow_blank=True)
                            wb[sheet].add_data_validation(dv)
                            dv.add(wb[sheet].cell(row=self.beg,column=3))

                    # Add units
                    if 'units' in schema_node:
                        wb[sheet].cell(row=self.beg,column=4).value = schema_node['units']

                    # Add required
                        # TODO: Make conditional formatting (e.g. red name if not entered)
                    if self.is_required():
                        wb[sheet].cell(row=self.beg,column=5).value = u'\u2713'  # Checkmark
                    wb[sheet].cell(row=self.beg,column=5).alignment = Alignment(horizontal='center')

                else:
                    # Not found in schema
                    comment = openpyxl.comments.Comment("Not found in schema.","ASHRAE 205")
                    wb[sheet].cell(row=self.beg,column=level_index).comment = comment
                    wb[sheet].cell(row=self.beg,column=level_index).font = Font(color='FF0001',bold=True)

                if self.sheet_ref:
                    wb[sheet].cell(row=self.beg,column=3).value = '$' + self.sheet_ref

                    # Hyperlink to referenced sheets
                    wb[sheet].cell(row=self.beg,column=3).hyperlink = f"#{self.sheet_ref}!A1"

                    if (self.child_sheet_type == SheetType.ARRAY and len(self.children) == 0):
                        # Make sheet for holding array values
                        array_sheet = unique_name_with_index(self.child_sheet,self.tree.sheets)
                        wb.create_sheet(array_sheet)
                        self.write_header(wb[array_sheet])

                        wb[array_sheet].cell(row=2,column=1).value = self.name
                        wb[array_sheet].cell(row=2,column=1).style = self.schema_style
                        wb[array_sheet].cell(row=3,column=1).style = self.schema_style

                        if schema_node:
                            if 'units' in schema_node:
                                wb[array_sheet].cell(row=3,column=1).value = schema_node['units']

                        row = 4
                        if self.value:
                            for value in self.value:
                                wb[array_sheet].cell(row=row,column=1).value = value
                                wb[array_sheet].cell(row=row,column=1).style = self.value_style
                                row += 1
                        else:
                            array_length = 5
                            if schema_node:
                                if 'maxItems' in schema_node:
                                    array_length = schema_node['maxItems']
                            for i in range(array_length):
                                wb[array_sheet].cell(row=row,column=1).style = self.value_style
                                row += 1

                if self.value is not None and self.sheet_ref is None:
                    wb[sheet].cell(row=self.beg,column=3).value = self.value

            # TODO: Something better here...a lot of repetition...
            elif self.sheet_type == SheetType.PERFORMANCE_MAP:
                if len(self.children) > 0:
                    level_index = 2
                else:
                    level_index = 3

                wb[sheet].cell(row=level_index,column=self.beg).value = self.name

                if '_variables' in self.parent.name:
                    if self.parent.name == 'grid_variables':
                        wb[sheet].cell(row=2,column=self.beg).style = self.grid_var_style
                        wb[sheet].cell(row=3,column=self.beg).style = self.grid_var_style
                        wb[sheet].cell(row=4,column=self.beg).style = self.grid_var_style
                        wb[sheet].cell(row=level_index,column=self.beg).alignment = Alignment(text_rotation=45)
                        if self.parent.grid_set:
                            row = 5
                            for value in self.parent.grid_set[self.name]:
                                wb[sheet].cell(row=row,column=self.beg).style = self.value_style
                                wb[sheet].cell(row=row,column=self.beg).value = value
                                row += 1
                        else:
                            # 2^n rows for spacing
                            row = 5
                            for i in range(2**(len(self.parent.children))):
                                wb[sheet].cell(row=row,column=self.beg).style = self.value_style
                                row += 1
                    else:
                        wb[sheet].cell(row=2,column=self.beg).style = self.schema_style
                        wb[sheet].cell(row=3,column=self.beg).style = self.schema_style
                        wb[sheet].cell(row=4,column=self.beg).style = self.schema_style
                        wb[sheet].cell(row=level_index,column=self.beg).alignment = Alignment(text_rotation=45)
                        if self.value is not None:
                            row = 5
                            for value in self.value:
                                wb[sheet].cell(row=row,column=self.beg).style = self.value_style
                                wb[sheet].cell(row=row,column=self.beg).value = value
                                row += 1
                        else:
                            # 2^n rows for spacing
                            row = 5
                            for i in range(2**(len(self.parent.parent.children[0].children))):
                                wb[sheet].cell(row=row,column=self.beg).style = self.value_style
                                row += 1

                if schema_node:
                    # Add units
                    if 'units' in schema_node:
                        wb[sheet].cell(row=4,column=self.beg).value = schema_node['units']

                    # Add required
                        # TODO: Make conditional formatting (e.g. red name if not entered)

                    # Add description
                    if 'description' in schema_node:
                        comment = openpyxl.comments.Comment(schema_node['description'],"ASHRAE 205")
                        wb[sheet].cell(row=level_index,column=self.beg).comment = comment

                else:
                    # Not found in schema
                    comment = openpyxl.comments.Comment("Not found in schema.","ASHRAE 205")
                    wb[sheet].cell(row=level_index,column=self.beg).comment = comment
                    wb[sheet].cell(row=level_index,column=self.beg).font = Font(color='FF0001',bold=True)

            # TODO: Something better here...a lot of repetition...
            elif self.sheet_type == SheetType.ARRAY:
                if len(self.children) > 0:
                    raise Exception("Were not handling nested items in an array yet!")

                wb[sheet].cell(row=2,column=self.beg).value = self.name
                wb[sheet].cell(row=2,column=self.beg).style = self.schema_style
                wb[sheet].cell(row=3,column=self.beg).style = self.schema_style

                row = 4
                if self.value is not None:
                    for value in self.value:
                        wb[sheet].cell(row=row,column=self.beg).value = value
                        wb[sheet].cell(row=row,column=self.beg).style = self.value_style

                        if schema_node:
                            # Enum validation
                            if 'enum' in schema_node:
                                enumerants = f'"{",".join(schema_node["enum"])}"'
                                if len(enumerants) < 256: # Apparent limitation of written lists (TODO: https://stackoverflow.com/a/33532984/1344457)
                                    dv = openpyxl.worksheet.datavalidation.DataValidation(type='list',formula1=enumerants,allow_blank=True)
                                    wb[sheet].add_data_validation(dv)
                                    dv.add(wb[sheet].cell(row=row,column=self.beg))

                        row += 1
                else:
                    array_length = 5
                    parent_schema_node = self.parent.get_schema_node()
                    if parent_schema_node:
                        if 'maxItems' in parent_schema_node:
                            array_length = parent_schema_node['maxItems']
                    for i in range(array_length):
                        wb[sheet].cell(row=row,column=self.beg).style = self.value_style
                        if schema_node:
                            # Enum validation
                            if 'enum' in schema_node:
                                enumerants = f'"{",".join(schema_node["enum"])}"'
                                if len(enumerants) < 256: # Apparent limitation of written lists (TODO: https://stackoverflow.com/a/33532984/1344457)
                                    dv = openpyxl.worksheet.datavalidation.DataValidation(type='list',formula1=enumerants,allow_blank=True)
                                    wb[sheet].add_data_validation(dv)
                                    dv.add(wb[sheet].cell(row=row,column=self.beg))
                        row += 1

                if schema_node:
                    # Add units
                    if 'units' in schema_node:
                        wb[sheet].cell(row=3,column=self.beg).value = schema_node['units']

                    # Add required
                        # TODO: Make conditional formatting (e.g. red name if not entered)

                    # Add description
                    if 'description' in schema_node:
                        comment = openpyxl.comments.Comment(schema_node['description'],"ASHRAE 205")
                        wb[sheet].cell(row=2,column=self.beg).comment = comment

                else:
                    # Not found in schema
                    comment = openpyxl.comments.Comment("Not found in schema.","ASHRAE 205")
                    wb[sheet].cell(row=2,column=self.beg).comment = comment
                    wb[sheet].cell(row=2,column=self.beg).font = Font(color='FF0001',bold=True)

        for child in self.children:
            child.write_node()

    def read_node(self):
        '''
        Translate XLSX content into nodes of a tree.
        '''
        ws = self.tree.workbook[self.child_sheet]
        end_node = False
        while not end_node:
            if self.child_sheet_type == SheetType.PERFORMANCE_MAP:
                # Everything from the perspective of parent node
                data_group = ws.cell(row=2,column=self.next_child_beg).value
                data_element = ws.cell(row=3,column=self.next_child_beg).value
                if data_group and data_group != self.name:
                    if data_group == 'grid_variables':
                        new_node = A205XLSXNode(data_group, parent=self)
                        new_node.add_grid_set({})
                        new_node.read_node()
                    elif data_group == 'lookup_variables':
                        if self.name == 'grid_variables':
                            # We hit the end of grid variables and need to conclude that node

                            # process grid set
                            grid_set = {}
                            for child in self.children:
                                grid_set[child.name] = child.value

                            self.add_grid_set(grid_set)

                            # check grid set?
                            grid_vars = process_grid_set(grid_set)

                            # reset grid variable values
                            for child in self.children:
                                child.value = grid_vars[child.name]

                            end_node = True
                        else:
                            new_node = A205XLSXNode(data_group, parent=self)
                            new_node.read_node()
                elif data_element:
                    if self.name not in ['grid_variables','lookup_variables']:
                        raise Exception(f"Invalid data group: '{self.name}'. Data groups in {self.parent.name} should be 'grid_variables' or 'lookup_variables'")
                    row = 5
                    end_of_column = False
                    value = []
                    while not end_of_column:
                        item = ws.cell(row=row,column=self.next_child_beg).value
                        if item is not None:
                            value.append(item)
                            row += 1
                        else:
                            end_of_column = True
                    new_node = A205XLSXNode(data_element, parent=self, value=value)
                else:
                    # End of sheet
                    end_node = True
            elif self.child_sheet_type == SheetType.ARRAY:
                # Everything from the perspective of parent node
                data_element = ws.cell(row=2,column=self.next_child_beg).value
                if data_element:
                    row = 4
                    end_of_column = False
                    value = []
                    while not end_of_column:
                        item = ws.cell(row=row,column=self.next_child_beg).value
                        if item is not None:
                            value.append(item)
                            row += 1
                        else:
                            end_of_column = True
                    new_node = A205XLSXNode(data_element, parent=self, value=value)
                else:
                    # End of sheet
                    end_node = True
            else:  # Flat Sheets
                data_group = ws.cell(row=self.next_child_beg,column=1).value
                data_element = ws.cell(row=self.next_child_beg,column=2).value
                cell_value = ws.cell(row=self.next_child_beg,column=3).value
                value = cell_value
                sheet_ref = None
                if type(cell_value) == str:
                    if cell_value[0] == '$':
                        value = None
                        sheet_ref = cell_value[1:]

                if data_group:
                    lineage = data_group.split(".")
                    if len(lineage) <= self.get_num_ancestors() + 1 and self.parent is not None:
                        # if lineage the same or shorter, this is not going to be a child node
                        end_node = True
                    else:
                        new_node = A205XLSXNode(lineage[-1], parent=self, value=value, sheet_ref=sheet_ref)
                        new_node.read_node()
                elif data_element:
                    if sheet_ref:
                        # Get array values from another sheet
                        row = 4
                        end_of_column = False
                        value = []
                        while not end_of_column:
                            item = self.tree.workbook[sheet_ref].cell(row=row,column=1).value
                            if item is not None:
                                value.append(item)
                                row += 1
                            else:
                                end_of_column = True
                    # Determine hierarchy level using number of spaces
                    level = (len(data_element) - len(data_element.lstrip(' ')))/self.white_space_multiplier
                    data_element = data_element.strip(' ')
                    generations = self.get_num_ancestors() - level
                    if generations > 0:
                        end_node = True
                    A205XLSXNode(data_element, parent=self.get_ancestor(generations), value=value, sheet_ref=sheet_ref)
                else:
                    # End of sheet
                    end_node = True

    def collect_content(self, content):
        '''
        Collect content from the tree and return it as a Python Dict.

        Used primarily to load contents from an XLSX sheet.
        '''
        if len(self.children) > 0:
            if self.child_sheet_type == SheetType.ARRAY:
                # Add list elements
                content[self.name] = []
                length = len(self.children[0].value)
                for i in range(length):
                    item = {}
                    for child in self.children:
                        if len(child.children) > 0:
                            raise Exception("Array sheets can only be one level deep!")
                        item[child.name] = child.value[i]
                    content[self.name].append(item)
            else:
                if self.name:
                    content[self.name] = {}
                    for child in self.children:
                        child.collect_content(content[self.name])
                else:
                    for child in self.children:
                        child.collect_content(content)
        else:
            content[self.name] = self.value


class A205XLSXTree:

    def __init__(self,schema_path=None):
        self.content = {}
        self.schema_type = ""
        if schema_path is None:
            schema_path = os.path.join(os.path.dirname(__file__),'..','schema-205',"build","schema","ASHRAE205.schema.json")
        self.schema = A205Schema(schema_path)
        self.root_node = None
        self.sheets = []
        self.template_args = {}
        self.template_args_used = {}

    def get_template_arg(self, arg):
        self.template_args_used[arg] = True
        return self.template_args[arg]

    def load_workbook(self, file_name):
        '''
        Create tree from XLSX workbook content
        '''
        self.workbook = openpyxl.load_workbook(file_name)
        # Find Primary RS worksheet
        rs_pattern = re.compile("^RS(\\d{4})$")
        for ws in self.workbook:
            if rs_pattern.match(ws.title):
                self.schema_type = ws.title

        # Load appropriate schema
        schema_path = os.path.join(os.path.dirname(__file__),'..','schema-205',"build","schema",f"{self.schema_type}.schema.json")
        self.schema = A205Schema(schema_path)

        self.root_node = A205XLSXNode(None, tree=self)
        self.root_node.read_node()
        return self

    def create_tree_from_content(self, content, parent):
        '''
        Create tree from Python Dict content
        '''
        if type(content) == list:
            # Transpose list content
            if len(content) > 0:
                # First make nodes for each of the elements in the array
                if type(content[0]) == dict:
                    self.create_tree_from_content(content[0], parent)
                    # Override values given based on first item with empty array
                    for child in parent.children:
                        child.value = []
                    for item in content:
                        for child in parent.children:
                            child.value.append(item[child.name])
        else:
            for item in content:
                if type(content[item]) == dict:
                    if "performance_map" in item:
                        sheet_ref = unique_name_with_index(item, self.sheets)
                    elif item[-len('_representation'):] == '_representation':
                        # Embedded rep
                        sheet_ref = unique_name_with_index(item, self.sheets)
                    else:
                        sheet_ref = None
                    new_node = A205XLSXNode(item, parent=parent, sheet_ref=sheet_ref)
                    if item == "grid_variables":
                        new_node.add_grid_set(self.schema.create_grid_set(self.content,new_node.lineage))
                    self.create_tree_from_content(content[item], new_node)
                elif type(content[item]) == list:
                    if len(content[item]) == 0:
                        # Create new sheet for blank array
                        sheet_ref = unique_name_with_index(item, self.sheets)
                        A205XLSXNode(item,parent=parent,sheet_ref=sheet_ref)
                    elif type(content[item][0]) == dict:
                        # Create new sheet for array
                        sheet_ref = unique_name_with_index(item, self.sheets)
                        new_node = A205XLSXNode(item,parent=parent,sheet_ref=sheet_ref)
                        self.create_tree_from_content(content[item], new_node)
                    else:
                        # plain array
                        if parent.sheet_type == SheetType.FLAT:
                            # simple array in it's own sheet
                            sheet_ref = unique_name_with_index(item, self.sheets)
                            new_node = A205XLSXNode(item,parent=parent,value=content[item],sheet_ref=sheet_ref)
                        else:
                            A205XLSXNode(item,parent=parent,value=content[item])
                else:
                    A205XLSXNode(item,parent=parent,value=content[item])

    def load(self, content):
        '''
        Create tree from Python Dict content
        '''
        self.content = content
        self.schema_type = content["metadata"]["schema"]

        # Load appropriate schema
        schema_path = os.path.join(os.path.dirname(__file__),'..','schema-205',"build","schema",f"{self.schema_type}.schema.json")
        self.schema = A205Schema(schema_path)

        self.root_node = A205XLSXNode(None, tree=self)
        self.create_tree_from_content(content, self.root_node)

    def create_tree_from_schema(self, node):
        '''
        Create an empty tree from the schema.

        Used largely for templating.
        '''
        schema_node = node.get_schema_node()

        # Handle alternative selectors
        if 'allOf' in schema_node:
            for alt in schema_node['allOf']:
                for selector_var, selector in alt['if']['properties'].items():
                    if selector_var in self.template_args:
                        if selector['const'] == self.get_template_arg(selector_var):
                            for property in alt['then']['properties']:
                                schema_node['properties'][property].update(alt['then']['properties'][property])
                    else:
                        raise Exception(f"No keyword arguments provided to determine template for '{node.name}'.")

        # typical nodes
        if 'properties' in schema_node:
            for item in schema_node['properties']:

                child_schema_node = self.schema.resolve(schema_node['properties'][item],step_in=False)

                # Typical cases
                option = None
                value = None
                sheet_ref = None

                # Special cases
                if item == 'schema_version':
                    value = self.schema.get_schema_version()
<<<<<<< HEAD
=======
                elif item == 'rs_id':
                    value = node.inner_rs
                elif item == 'rs_instance':
                    option = get_rs_index(node.inner_rs)
>>>>>>> c0ce73421ed7dd8ab84c20c849d4785e4954f974
                elif 'performance_map' == item[:len('performance_map')]:
                    sheet_ref = unique_name_with_index(item, self.sheets)
                elif 'items' in child_schema_node and node.sheet_type == SheetType.FLAT:
                    sheet_ref = unique_name_with_index(item, self.sheets)
                elif item[-len('_representation'):] == '_representation':
                    # Embedded rep spec
                    sheet_ref = item
                elif item in self.template_args:
                    # General keyword value setting
                    value = self.get_template_arg(item)

                self.create_tree_from_schema(A205XLSXNode(item, parent=node, value=value, option=option, sheet_ref=sheet_ref))

        # List nodes:
        if 'items' in schema_node:
            schema_node = self.schema.resolve(schema_node['items'],step_in=False)
            if 'properties' in schema_node:
                for item in schema_node['properties']:
                    self.create_tree_from_schema(A205XLSXNode(item, parent=node))

        # oneOf nodes
        if 'oneOf' in schema_node:
<<<<<<< HEAD
            raise Exception(f"oneOf not yet handled for '{node.name}'.")
=======
            if node.inner_rs == 'RS0003' and node.name == 'performance_map' and 'operation_speed_control_type' in self.template_args:
                template_arg_value = self.get_template_arg('operation_speed_control_type')
                if template_arg_value == 'CONTINUOUS':
                    schema_node = self.schema.resolve(schema_node['oneOf'][0],step_in=False)
                    node.options[-1] = 0
                elif template_arg_value == 'DISCRETE':
                    schema_node = self.schema.resolve(schema_node['oneOf'][1],step_in=False)
                    node.options[-1] = 1
                else:
                    raise Exception(f"Invalid 'performance_map_type': {template_arg_value}")

                for item in schema_node['properties']:
                    self.create_tree_from_schema(A205XLSXNode(item, parent=node))
            else:
                raise Exception(f"No keyword arguments provided to determine template for '{node.name}'.")
>>>>>>> c0ce73421ed7dd8ab84c20c849d4785e4954f974

    def template_tree(self, repspec, **kwargs):
        '''
        Generate empty tree content based on the schema for a specific RS?
        Generate an XLSX template based on the schema for a specific RS.

        kwargs:
          any data element and value in the schema
        '''
        self.schema_type = repspec
        self.template_args = kwargs
        for arg in self.template_args:
            self.template_args_used[arg] = False
        self.root_node = A205XLSXNode(None, tree=self)
        self.create_tree_from_schema(self.root_node)
<<<<<<< HEAD
        for arg in self.template_args_used:
            if not self.template_args_used[arg]:
                raise Exception(f"{self.schema_type} unused template argument: \"{arg}\". ")
=======
        #for arg in self.template_args_used:
        #    if not self.template_args_used[arg]:
        #        raise Exception(f"Unused template argument: \"{arg}\".")
>>>>>>> c0ce73421ed7dd8ab84c20c849d4785e4954f974

    def template(self, repspec, output_path, **kwargs):
        '''
        Generate empty tree content based on the schema for a specific RS?
        Generate an XLSX template based on the schema for a specific RS.
        '''
        self.template_tree(repspec, **kwargs)

        output_path = os.path.join(output_path)
        self.save(output_path)

    def save(self, file_name):
        '''
        Save tree as workbook
        '''
        self.workbook = openpyxl.Workbook()

        # Write tree content
        self.workbook.active.title = self.schema_type
        self.root_node.write_header(self.workbook[self.schema_type])
        self.root_node.write_node()

        self.workbook.save(file_name)

    def get_content(self):
        '''
        returns tree content
        '''
        content = {}
        self.root_node.collect_content(content)
        return content

def template(repspec, output_path, **kwargs):
    '''
    Generate an XLSX template based on the schema for a specific RS
    '''
    # Load appropriate schema
    schema_path = os.path.join(os.path.dirname(__file__),'..','schema-205',"build","schema",f"{repspec}.schema.json")
    tree = A205XLSXTree(schema_path=schema_path)
    tree.template(repspec, output_path, **kwargs)

def generate_templates(output_dir, config):
    for rs, templates in config.items():
        for t in templates:
            file_name_components = [rs]
            if t["file-name-suffix"]:
                file_name_components.append(t["file-name-suffix"])
            file_name_components.append("template.a205.xlsx")
            file_name = '-'.join(file_name_components)
            template(rs,os.path.join(output_dir,file_name), **t["keywords"])
