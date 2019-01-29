import os
import sys
import json
import re
import pandas as pd
from openpyxl import load_workbook
from subprocess import call, check_output

if sys.argv[1] == '-f':
    build_dir = '../../build'
    source_dir = '../..'
else:
    build_dir = sys.argv[1]
    source_dir = sys.argv[2]

sys.path.append(source_dir + '/vendor/FlatBuffers/python')
sys.path.append(build_dir + '/python' )

import flatbuffers
import ASHRAE205

# TODO import below as needed?
import RS
import CompressorType
import FluidType
import RefrigerantType
import RS0001.RS0001_Root
import RS0001.GeneralData
import RS0001.DriveType
import RS0001.AHRIRatings
import RS0001.StandardRatingVersion
import RS0001.PerformanceMapCooling
import RS0001.PerformancePointCooling
import RS0001.PerformanceMapStandby
import RS0001.PerformancePointStandby

def upperfirst(x):
    return x[0].upper() + x[1:]

source_xlsx = source_dir + '/examples/python-writer/chiller-example.xlsx'

worksheet = load_workbook(source_xlsx)

# Check Equipment type
RSID = pd.read_excel(source_xlsx,'ASHRAE205').iloc[0].RSID

with open(build_dir + '/schema/ASHRAE205.json') as schema_file:
    schema = json.load(schema_file)

builder = flatbuffers.Builder(1024)

def setup_table(table_name):

    # Check table name validity
    if table_name in worksheet.sheetnames:
        table = pd.read_excel(source_xlsx,table_name)
    else:
        print("Table, {0}, not found in {1}".format(table_name,source_xlsx))

    for i in range(0,len(schema["objects"])):
        if schema["objects"][i]["name"] == RSID + '.' + table_name:
            table_index = i
            break

    elements = []

    for field in schema["objects"][table_index]["fields"]:
        name = field["name"]
        enum = None
        enum_options = []
        if "index" in field["type"]:
            tp = 'Enum'
            enum = schema["enums"][field["type"]["index"]]["name"]
            for opt in schema["enums"][field["type"]["index"]]["values"]:
                enum_options.append(opt["name"])
        else:
            tp = field["type"]["base_type"]

        required = True
        for att in field["attributes"]:
            if att["key"] == "optional":
                required = False
                break


        if required and name not in list(table):
            raise Exception("Required field, '{0}', not found in spreadsheet.".format(name))
        if tp == 'String':
            value = builder.CreateString(str(table.iloc[0][name]))
        elif tp == 'Double':
            value = float(table.iloc[0][name])
        elif tp == 'Bool':
            value = bool(table.iloc[0][name])
        elif tp == 'Enum':
            if RSID in enum:
                enum_name = re.match(RSID + "\.(.*)",enum).group(1)
            else:
                enum_name = enum
            if table.iloc[0][name] not in enum_options:
                enum_list = ""
                for opt in enum_options:
                    enum_list += "\n  - " + opt
                raise Exception("Invalid enumeration value for {0}, '{1}'. Valid options are: {2}".format(enum_name,table.iloc[0][name],enum_list))
            value = eval('{0}.{1}().{2}'.format(enum,enum_name,table.iloc[0][name]))
        else:
            print("Unknown data type: '{}'".format(tp))
            value = table.iloc[0][name]

        elements.append({"name":name, "type":tp, "value":value})

    # populate builder
    eval("{0}.{1}.{1}Start(builder)".format(RSID,table_name))
    for e in elements:
        eval("{0}.{1}.{1}Add{2}(builder,{3})".format(RSID,table_name,upperfirst(e["name"]),e["value"]))
    table_buffer = eval("{0}.{1}.{1}End(builder)".format(RSID,table_name))
    return table_buffer

gen_data = setup_table('GeneralData')
ahri_ratings = setup_table('AHRIRatings')

# Build PerformanceMapCooling table
x_PMC = pd.read_excel(source_xlsx,'PerformanceMapCooling')

RS0001.PerformanceMapCooling.PerformanceMapCoolingStartPerformancePointsCoolingVector(builder,len(x_PMC))
for ind, point in x_PMC.iterrows():
    RS0001.PerformancePointCooling.CreatePerformancePointCooling(builder,point.evaporatorFluidVolumetricFlowRate,point.evaporatorFluidLeavingTemperature,point.condenserFluidVolumetricFlowRate,point.condenserFluidEnteringTemperature,point.netRefrigeratingCapacityFraction,point.inputPower,point.netRefrigeratingCapacity,point.heatLossFraction,point.evaporatorFluidEnteringTemperature,point.condenserFluidLeavingTemperature,point.evaporatorFluidDifferentialPressure,point.condenserFluidDifferentialPressure,int(point.numberCompressorsOnline))
cool_points = builder.EndVector(len(x_PMC))

RS0001.PerformanceMapCooling.PerformanceMapCoolingStart(builder)
RS0001.PerformanceMapCooling.PerformanceMapCoolingAddPerformancePointsCooling(builder,cool_points)
cool_map = RS0001.PerformanceMapCooling.PerformanceMapCoolingEnd(builder)

# Build PerformanceMapStandby table
x_PMS = pd.read_excel(source_xlsx,'PerformanceMapStandby')

RS0001.PerformanceMapStandby.PerformanceMapStandbyStartPerformancePointsStandbyVector(builder,len(x_PMS))
for ind, point in x_PMS.iterrows():
    RS0001.PerformancePointStandby.CreatePerformancePointStandby(builder,point.chillerEnvironmentDryBulbTemperature,point.inputPower)
sb_points = builder.EndVector(len(x_PMS))

RS0001.PerformanceMapStandby.PerformanceMapStandbyStart(builder)
RS0001.PerformanceMapStandby.PerformanceMapStandbyAddPerformancePointsStandby(builder,sb_points)
sb_map = RS0001.PerformanceMapStandby.PerformanceMapStandbyEnd(builder)

# Add all tables to RS0001 root table
RS0001.RS0001_Root.RS0001_RootStart(builder)
RS0001.RS0001_Root.RS0001_RootAddGeneralDataTable(builder,gen_data)
RS0001.RS0001_Root.RS0001_RootAddAHRIRatingsTable(builder,ahri_ratings)
RS0001.RS0001_Root.RS0001_RootAddPerformanceMapCoolingTable(builder,cool_map)
RS0001.RS0001_Root.RS0001_RootAddPerformanceMapStandbyTable(builder,sb_map)
chiller = RS0001.RS0001_Root.RS0001_RootEnd(builder)

# Build ASHRAE205 table
x_ASHRAE205 = pd.read_excel(source_xlsx,'ASHRAE205').iloc[0]

title = builder.CreateString(x_ASHRAE205.RSTitle)
version = builder.CreateString(x_ASHRAE205.RSVersion)
ID = builder.CreateString(x_ASHRAE205.ID)
description = builder.CreateString(x_ASHRAE205.description)
timestamp = builder.CreateString(x_ASHRAE205.dataTimestamp)

ASHRAE205.ASHRAE205Start(builder)
ASHRAE205.ASHRAE205AddRSInstanceType(builder,RS.RS().RS0001_RS0001_Root)
ASHRAE205.ASHRAE205AddRSInstance(builder,chiller)
ASHRAE205.ASHRAE205AddRSTitle(builder,title)
ASHRAE205.ASHRAE205AddRSVersion(builder,version)
ASHRAE205.ASHRAE205AddID(builder,ID)
ASHRAE205.ASHRAE205AddDataTimestamp(builder,timestamp)
ASHRAE205.ASHRAE205AddDescription(builder,description)
a205 = ASHRAE205.ASHRAE205End(builder)


# Complete buffer
builder.Finish(a205,)
buf = builder.Output()

# Write flatbuffer
flatbuff_path = build_dir + '/examples/python-writer'
flatbuff_name = 'chiller.a205'

if not os.path.exists(flatbuff_path):
    os.makedirs(flatbuff_path)
with open(flatbuff_path + '/' + flatbuff_name,'wb') as bfile:
    bfile.write(buf)

# Write equivalent JSON
call('{0}/vendor/FlatBuffers/flatc --json --strict-json --raw-binary --defaults-json -o {1} {2}/schema/ASHRAE205.fbs -- {1}/{3}'.format(build_dir,flatbuff_path,source_dir,flatbuff_name),shell=True)

# testing
chiller_read = ASHRAE205.ASHRAE205.GetRootAsASHRAE205(buf,0)

chiller_read.RSTitle()

rs_type = chiller_read.RSInstanceType()
rs = RS0001.RS0001_Root.RS0001_Root()
rs.Init(chiller_read.RSInstance().Bytes, chiller_read.RSInstance().Pos)

rs.GeneralDataTable().Manufacturer()
